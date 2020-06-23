import torch
import torch.nn.functional as F
import argparse
# import xlsxwriter
# import xlrd
import openpyxl

from transformers import (AlbertConfig,
                          AlbertForSequenceClassification,
                          AlbertTokenizer,
                          )


class SentimentAnalyzer:
    def __init__(self, path, model_type):
        self.path = path
        self.model_type = model_type
        self.tokenizer = AlbertTokenizer.from_pretrained(self.model_type, do_lower_case=True)
        self.model = AlbertForSequenceClassification.from_pretrained(self.path)
        self.device = "cuda" if torch.cuda.is_available() else "cpu"
        self.model.to(self.device)
        self.model.eval()

    def convert_to_features(self, sentence):
        text_a = sentence
        text_b = None
        max_length = 512
        pad_on_left = False
        pad_token = self.tokenizer.convert_tokens_to_ids([self.tokenizer.pad_token])[0]
        pad_token_segment_id = 0
        mask_padding_with_zero = True

        inputs = self.tokenizer.encode_plus(
            text_a,
            text_b,
            add_special_tokens=True,
            max_length=max_length)
        input_ids, token_type_ids = inputs["input_ids"], inputs["token_type_ids"]

        # The mask has 1 for real tokens and 0 for padding tokens. Only real
        # tokens are attended to.
        attention_mask = [1 if mask_padding_with_zero else 0] * len(input_ids)

        # Zero-pad up to the sequence length.
        padding_length = max_length - len(input_ids)
        if pad_on_left:
            input_ids = ([pad_token] * padding_length) + input_ids
            attention_mask = ([0 if mask_padding_with_zero else 1] * padding_length) + attention_mask
            token_type_ids = ([pad_token_segment_id] * padding_length) + token_type_ids
        else:
            input_ids = input_ids + ([pad_token] * padding_length)
            attention_mask = attention_mask + ([0 if mask_padding_with_zero else 1] * padding_length)
            token_type_ids = token_type_ids + ([pad_token_segment_id] * padding_length)
        return [input_ids, attention_mask, token_type_ids]

    def convert_to_tensors(self, features):
        input_ids = torch.tensor([features[0]],
                                 dtype=torch.long)

        attention_mask = torch.tensor([features[1]],
                                      dtype=torch.long)

        token_type_ids = torch.tensor([features[2]],
                                      dtype=torch.long)

        inputs = {'input_ids': input_ids,
                  'attention_mask': attention_mask}
        return inputs

    def interpret_result(self, output):
        result = {}
        logits = F.softmax(output[0][0], dim=0)
        logits_label = torch.argmax(logits, dim=0)
        logits_label = logits_label.detach().cpu().numpy().tolist()
        score = round(logits[logits_label].detach().cpu().numpy().tolist(), 5)
        logits = logits.detach().cpu().numpy().tolist()
        logits = [round(logit, 4) for logit in logits]
        result['label'] = logits_label
        result['confidence'] = score
        result['logits'] = logits
        return result

    def predict(self, text):
        features = self.convert_to_features(text)
        tensor = self.convert_to_tensors(features)
        outputs = self.model(**tensor)
        result = self.interpret_result(outputs)
        return result


if __name__ ==  '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument("--path", required=True, type=str, help="path to finetuned albert model")
    ap.add_argument("--model_name", required=True, type=str, help="variant of albert to use")
    ap.add_argument("--file", required=False, type=str, help="path to xlsx file to feed in for predictions")
    args = vars(ap.parse_args())
    analyzer = SentimentAnalyzer(args['path'], args['model_name'])
    if (args['file']):
        print("----file present: "+  args['file'] + "----")
        wkbk = openpyxl.load_workbook(args['file'])
        sheet = wkbk.active # get working sheet
        num_entries = sheet.max_row
        for i in range(1, num_entries+1):
            entry = sheet.cell(row = i, column = 1).value
            print(entry)
            prediction = analyzer.predict(entry)
            predict_label = prediction['label']
            print(f"sentiment prediction: {predict_label} w confidence: {prediction['confidence']}")
            print("\n")
            output = sheet.cell(row = i, column = 3)
            output.value = predict_label
        wkbk.save(args['file'])

    else:
        # text = "Mysterious coronavirus condition ‘happy hypoxia’ baffles doctors"
        text = 'trump work lets reopen'
        print(text)
        print(analyzer.predict(text))
